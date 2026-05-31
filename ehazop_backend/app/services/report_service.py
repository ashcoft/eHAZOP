"""Report generation service for PDF and Excel exports."""

import io
from datetime import datetime
from typing import Any

from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from ehazop_backend.app.models.user import Study
from ehazop_backend.app.models.hazard import Node, Deviation, Cause, Consequence, Safeguard, RiskRanking
from ehazop_backend.app.models.action import Recommendation
from ehazop_backend.app.services.storage_service import StorageService


class ReportService:
    """Service for generating study reports in PDF and Excel formats."""

    def __init__(self, db: AsyncSession):
        self.db = db
        self.storage = StorageService(db)

    async def generate_pdf_report(
        self,
        study_id: str,
        include_safeguards: bool = True,
        include_actions: bool = True,
    ) -> bytes:
        """Generate a PDF report for a study."""
        study = await self._get_study_with_data(study_id)
        if not study:
            raise ValueError(f"Study {study_id} not found")

        # Generate HTML content
        html_content = self._generate_html_report(study, include_safeguards, include_actions)

        # Convert to PDF using WeasyPrint
        try:
            from weasyprint import HTML
            pdf_bytes = HTML(string=html_content).write_pdf()
            return pdf_bytes
        except ImportError:
            raise RuntimeError("WeasyPrint is required for PDF generation")

    async def generate_excel_report(
        self,
        study_id: str,
    ) -> bytes:
        """Generate an Excel report for a study."""
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
        except ImportError:
            raise RuntimeError("openpyxl is required for Excel generation")

        study = await self._get_study_with_data(study_id)
        if not study:
            raise ValueError(f"Study {study_id} not found")

        wb = openpyxl.Workbook()

        # Create sheets
        self._create_summary_sheet(wb, study)
        self._create_worksheet_sheet(wb, study)
        self._create_actions_sheet(wb, study)

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()

    async def _get_study_with_data(self, study_id: str) -> dict[str, Any] | None:
        """Get study with all related data."""
        result = await self.db.execute(select(Study).where(Study.id == study_id))
        study = result.scalar_one_or_none()
        if not study:
            return None

        # Get nodes with deviations
        nodes_result = await self.db.execute(
            select(Node).where(Node.study_id == study_id, Node.is_active == True)
        )
        nodes = list(nodes_result.scalars().all())

        study_data = {
            "id": study.id,
            "name": study.name,
            "study_type": study.study_type,
            "facility": study.facility,
            "status": study.status,
            "revision": study.revision,
            "description": study.description,
            "created_at": study.created_at.isoformat() if study.created_at else None,
            "completed_at": study.completed_at.isoformat() if study.completed_at else None,
            "nodes": [],
        }

        for node in nodes:
            deviations_result = await self.db.execute(
                select(Deviation).where(Deviation.node_id == node.id)
            )
            deviations = list(deviations_result.scalars().all())

            node_data = {
                "reference": node.reference,
                "name": node.name,
                "design_intent": node.design_intent,
                "equipment_type": node.equipment_type,
                "deviations": [],
            }

            for deviation in deviations:
                # Get causes, consequences, safeguards
                causes_result = await self.db.execute(
                    select(Cause).where(Cause.deviation_id == deviation.id)
                )
                consequences_result = await self.db.execute(
                    select(Consequence).where(Consequence.deviation_id == deviation.id)
                )
                safeguards_result = await self.db.execute(
                    select(Safeguard).where(Safeguard.deviation_id == deviation.id)
                )
                risk_result = await self.db.execute(
                    select(RiskRanking).where(RiskRanking.deviation_id == deviation.id)
                )
                actions_result = await self.db.execute(
                    select(Recommendation).where(Recommendation.deviation_id == deviation.id)
                )

                node_data["deviations"].append({
                    "reference": deviation.reference,
                    "deviation_text": deviation.deviation_text,
                    "location": deviation.location,
                    "status": deviation.status,
                    "causes": [c.description for c in causes_result.scalars().all()],
                    "consequences": [c.description for c in consequences_result.scalars().all()],
                    "safeguards": [s.description for s in safeguards_result.scalars().all()],
                    "risk_rankings": [
                        {
                            "category": r.category,
                            "severity": r.severity,
                            "likelihood": r.likelihood,
                            "risk_score": r.risk_score,
                            "risk_band": r.risk_band,
                        }
                        for r in risk_result.scalars().all()
                    ],
                    "recommendations": [
                        {
                            "reference": a.reference,
                            "description": a.description,
                            "priority": a.priority,
                            "status": a.status,
                            "target_date": a.target_date.isoformat() if a.target_date else None,
                        }
                        for a in actions_result.scalars().all()
                    ],
                })

            study_data["nodes"].append(node_data)

        return study_data

    def _generate_html_report(
        self,
        study: dict[str, Any],
        include_safeguards: bool,
        include_actions: bool,
    ) -> str:
        """Generate HTML report content."""
        # Import Jinja2 for templating
        try:
            from jinja2 import Template
        except ImportError:
            # Simple string replacement if Jinja2 not available
            return self._generate_simple_html(study)

        template = """
<!DOCTYPE html>
<html>
<head>
    <meta charset="UTF-8">
    <title>{{ study.name }} - {{ study.study_type }} Report</title>
    <style>
        body { font-family: Arial, sans-serif; margin: 40px; }
        h1 { color: #333; border-bottom: 2px solid #0066cc; padding-bottom: 10px; }
        h2 { color: #0066cc; margin-top: 30px; }
        .meta { background: #f5f5f5; padding: 15px; margin: 20px 0; border-radius: 5px; }
        .meta-item { margin: 5px 0; }
        .node { margin: 20px 0; padding: 15px; border: 1px solid #ddd; border-radius: 5px; }
        .deviation { margin: 15px 0; padding: 10px; background: #fafafa; border-left: 3px solid #0066cc; }
        .risk-high { color: #d32f2f; font-weight: bold; }
        .risk-medium { color: #f57c00; font-weight: bold; }
        .risk-low { color: #388e3c; }
        table { width: 100%; border-collapse: collapse; margin: 10px 0; }
        th, td { border: 1px solid #ddd; padding: 8px; text-align: left; }
        th { background: #0066cc; color: white; }
        .footer { margin-top: 40px; padding-top: 20px; border-top: 1px solid #ddd; font-size: 12px; color: #666; }
    </style>
</head>
<body>
    <h1>{{ study.name }}</h1>
    <div class="meta">
        <div class="meta-item"><strong>Study Type:</strong> {{ study.study_type }}</div>
        <div class="meta-item"><strong>Facility:</strong> {{ study.facility }}</div>
        <div class="meta-item"><strong>Status:</strong> {{ study.status }}</div>
        <div class="meta-item"><strong>Revision:</strong> {{ study.revision }}</div>
        <div class="meta-item"><strong>Date:</strong> {{ study.created_at }}</div>
    </div>
    
    {% if study.description %}
    <h2>Description</h2>
    <p>{{ study.description }}</p>
    {% endif %}
    
    <h2>Nodes and Deviations</h2>
    {% for node in study.nodes %}
    <div class="node">
        <h3>{{ node.reference }}: {{ node.name }}</h3>
        {% if node.design_intent %}
        <p><strong>Design Intent:</strong> {{ node.design_intent }}</p>
        {% endif %}
        
        {% for deviation in node.deviations %}
        <div class="deviation">
            <strong>{{ deviation.reference }}</strong>: {{ deviation.deviation_text or 'N/A' }}
            <br><em>Status: {{ deviation.status }}</em>
            
            {% if deviation.risk_rankings %}
            <p><strong>Risk Assessment:</strong>
            {% for risk in deviation.risk_rankings %}
                <span class="risk-{{ 'high' if risk.risk_band == 'Very High' or risk.risk_band == 'High' else 'medium' if risk.risk_band == 'Medium' else 'low' }}">
                    {{ risk.category }}: {{ risk.risk_score }} ({{ risk.risk_band }})
                </span>
            {% endfor %}
            </p>
            {% endif %}
            
            {% if deviation.causes %}
            <p><strong>Causes:</strong></p>
            <ul>
            {% for cause in deviation.causes %}
                <li>{{ cause }}</li>
            {% endfor %}
            </ul>
            {% endif %}
            
            {% if deviation.consequences %}
            <p><strong>Consequences:</strong></p>
            <ul>
            {% for consequence in deviation.consequences %}
                <li>{{ consequence }}</li>
            {% endfor %}
            </ul>
            {% endif %}
            
            {% if include_safeguards and deviation.safeguards %}
            <p><strong>Safeguards:</strong></p>
            <ul>
            {% for safeguard in deviation.safeguards %}
                <li>{{ safeguard }}</li>
            {% endfor %}
            </ul>
            {% endif %}
            
            {% if include_actions and deviation.recommendations %}
            <p><strong>Recommendations:</strong></p>
            <ul>
            {% for action in deviation.recommendations %}
                <li>{{ action.reference }}: {{ action.description }} ({{ action.status }})</li>
            {% endfor %}
            </ul>
            {% endif %}
        </div>
        {% endfor %}
    </div>
    {% endfor %}
    
    <div class="footer">
        <p>Generated: {{ now }} | EHAZOP Platform</p>
    </div>
</body>
</html>
        """
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        tmpl = Template(template)
        return tmpl.render(study=study, now=now, include_safeguards=include_safeguards, include_actions=include_actions)

    def _generate_simple_html(self, study: dict[str, Any]) -> str:
        """Generate simple HTML without Jinja2."""
        html = f"""
<!DOCTYPE html>
<html>
<head>
    <meta charset="UTF-8">
    <title>{study['name']} Report</title>
</head>
<body>
    <h1>{study['name']}</h1>
    <p><strong>Type:</strong> {study['study_type']} | <strong>Facility:</strong> {study['facility']} | <strong>Status:</strong> {study['status']}</p>
    <hr>
"""
        for node in study.get("nodes", []):
            html += f"<h2>{node['reference']}: {node['name']}</h2>"
            for deviation in node.get("deviations", []):
                html += f"<div style='margin-left:20px;padding:10px;border-left:3px solid #0066cc;'>"
                html += f"<strong>{deviation['reference']}</strong>: {deviation.get('deviation_text', 'N/A')}"
                if deviation.get("risk_rankings"):
                    risks = ", ".join([f"{r['category']}: {r['risk_score']} ({r['risk_band']})" for r in deviation['risk_rankings']])
                    html += f"<br><em>Risk: {risks}</em>"
                html += "</div>"
        html += "</body></html>"
        return html

    def _create_summary_sheet(self, wb: Any, study: dict[str, Any]) -> None:
        """Create summary sheet in Excel workbook."""
        ws = wb.active
        ws.title = "Summary"

        ws["A1"] = "Study Report Summary"
        ws["A1"].font = Font(bold=True, size=16)
        
        ws["A3"] = "Study Name"
        ws["B3"] = study["name"]
        ws["A4"] = "Study Type"
        ws["B4"] = study["study_type"]
        ws["A5"] = "Facility"
        ws["B5"] = study["facility"]
        ws["A6"] = "Status"
        ws["B6"] = study["status"]
        ws["A7"] = "Revision"
        ws["B7"] = study["revision"]

    def _create_worksheet_sheet(self, wb: Any, study: dict[str, Any]) -> None:
        """Create worksheet sheet with deviations."""
        ws = wb.create_sheet("Worksheet")

        # Headers
        headers = ["Node Ref", "Node Name", "Dev Ref", "Deviation", "Location", "Status", 
                   "Causes", "Consequences", "Safeguards", "Risk", "Recommendations"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="0066cc", end_color="0066cc", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")

        row = 2
        for node in study.get("nodes", []):
            for deviation in node.get("deviations", []):
                risk_str = ", ".join([
                    f"{r['category']}: {r['risk_score']}" 
                    for r in deviation.get("risk_rankings", [])
                ])
                causes_str = "; ".join(deviation.get("causes", []))
                consequences_str = "; ".join(deviation.get("consequences", []))
                safeguards_str = "; ".join(deviation.get("safeguards", []))
                actions_str = "; ".join([
                    f"{a['reference']}: {a['description']}" 
                    for a in deviation.get("recommendations", [])
                ])

                ws.cell(row=row, column=1).value = node["reference"]
                ws.cell(row=row, column=2).value = node["name"]
                ws.cell(row=row, column=3).value = deviation["reference"]
                ws.cell(row=row, column=4).value = deviation.get("deviation_text", "")
                ws.cell(row=row, column=5).value = deviation.get("location", "")
                ws.cell(row=row, column=6).value = deviation["status"]
                ws.cell(row=row, column=7).value = causes_str
                ws.cell(row=row, column=8).value = consequences_str
                ws.cell(row=row, column=9).value = safeguards_str
                ws.cell(row=row, column=10).value = risk_str
                ws.cell(row=row, column=11).value = actions_str
                row += 1

    def _create_actions_sheet(self, wb: Any, study: dict[str, Any]) -> None:
        """Create actions sheet with recommendations."""
        ws = wb.create_sheet("Actions")

        # Headers
        headers = ["Reference", "Description", "Priority", "Status", "Target Date"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="0066cc", end_color="0066cc", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")

        row = 2
        for node in study.get("nodes", []):
            for deviation in node.get("deviations", []):
                for action in deviation.get("recommendations", []):
                    ws.cell(row=row, column=1).value = action["reference"]
                    ws.cell(row=row, column=2).value = action["description"]
                    ws.cell(row=row, column=3).value = action["priority"]
                    ws.cell(row=row, column=4).value = action["status"]
                    ws.cell(row=row, column=5).value = action.get("target_date", "")
                    row += 1

    async def save_report(
        self,
        study_id: str,
        report_type: str,  # pdf, excel
        content: bytes,
        filename: str,
    ) -> dict[str, Any]:
        """Save a generated report to storage."""
        return await self.storage.upload_file(
            content=content,
            filename=filename,
            file_type=report_type,
            study_id=study_id,
        )